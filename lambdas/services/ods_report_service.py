import os
import tempfile
from datetime import datetime

from enums.dynamo_filter import AttributeOperator
from enums.file_type import FileType
from enums.lambda_error import LambdaError
from enums.metadata_field_names import DocumentReferenceMetadataFields
from enums.patient_ods_inactive_status import PatientOdsInactiveStatus
from enums.repository_role import RepositoryRole
from openpyxl.workbook import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from services.base.dynamo_service import DynamoDBService
from services.base.s3_service import S3Service
from utils.audit_logging_setup import LoggingService
from utils.dynamo_query_filter_builder import DynamoQueryFilterBuilder
from utils.lambda_exceptions import OdsReportException
from utils.request_context import request_context

logger = LoggingService(__name__)


class OdsReportService:
    def __init__(self):
        download_report_aws_role_arn = os.getenv("PRESIGNED_ASSUME_ROLE")
        self.s3_service = S3Service(custom_aws_role=download_report_aws_role_arn)
        self.dynamo_service = DynamoDBService()
        self.table_name = os.getenv("LLOYD_GEORGE_DYNAMODB_NAME")
        self.reports_bucket = os.getenv("STATISTICAL_REPORTS_BUCKET")
        self.temp_output_dir = ""

    def get_nhs_numbers_by_ods(
        self,
        ods_code: str,
        is_pre_signed_needed: bool = False,
        is_upload_to_s3_needed: bool = False,
        file_type_output: FileType = FileType.CSV,
    ):
        results = self.scan_table_with_filter(ods_code)
        nhs_numbers = {
            item.get(DocumentReferenceMetadataFields.NHS_NUMBER.value)
            for item in results
            if item.get(DocumentReferenceMetadataFields.NHS_NUMBER.value)
        }
        if is_upload_to_s3_needed:
            self.temp_output_dir = tempfile.mkdtemp()
        return self.create_and_save_ods_report(
            ods_code,
            nhs_numbers,
            is_pre_signed_needed,
            is_upload_to_s3_needed,
            file_type_output,
        )

    def scan_table_with_filter(self, ods_code: str):
        ods_codes = [ods_code]
        authorization_user = getattr(request_context, "authorization", {})
        if (
            authorization_user
            and authorization_user.get("repository_role") == RepositoryRole.PCSE.value
        ):
            ods_codes = [
                PatientOdsInactiveStatus.SUSPENDED,
                PatientOdsInactiveStatus.DECEASED,
            ]
        ods_filter_expression = self.build_filter_expression(ods_codes)
        field_names_expression = ",".join(DocumentReferenceMetadataFields.list())

        results = []
        response = self.dynamo_service.scan_whole_table(
            table_name=self.table_name,
            project_expression=field_names_expression,
            filter_expression=ods_filter_expression,
        )
        results.extend(response)
        if not results:
            logger.info("No records found for ODS code {}".format(ods_code))
            raise OdsReportException(404, LambdaError.NoDataFound)

        return results

    @staticmethod
    def build_filter_expression(ods_codes: list[str]):
        filter_builder = DynamoQueryFilterBuilder()
        return (
            filter_builder.add_condition(
                attribute=DocumentReferenceMetadataFields.CURRENT_GP_ODS.value,
                attr_operator=AttributeOperator.IN,
                filter_value=ods_codes,
            )
            .add_condition(
                attribute=str(DocumentReferenceMetadataFields.DELETED.value),
                attr_operator=AttributeOperator.EQUAL,
                filter_value="",
            )
            .build()
        )

    def query_table_by_index(self, ods_code: str):
        results = []

        response = self.dynamo_service.query_table_by_index(
            table_name=self.table_name,
            index_name="OdsCodeIndex",
            search_key=DocumentReferenceMetadataFields.CURRENT_GP_ODS.value,
            search_condition=ods_code,
        )
        results += response["Items"]

        if not response["Items"]:
            logger.info("No records found for ODS code {}".format(ods_code))
        while "LastEvaluatedKey" in response:
            response = self.dynamo_service.query_table_by_index(
                table_name=self.table_name,
                index_name="OdsCodeIndex",
                exclusive_start_key=response["LastEvaluatedKey"],
                search_key=DocumentReferenceMetadataFields.CURRENT_GP_ODS.value,
                search_condition=ods_code,
            )
            results += response["Items"]
        return results

    def create_and_save_ods_report(
        self,
        ods_code: str,
        nhs_numbers: set[str],
        create_pre_signed_url: bool = False,
        upload_to_s3: bool = False,
        file_type_output: FileType = FileType.CSV,
    ):
        now = datetime.now()
        formatted_time = now.strftime("%Y-%m-%d_%H-%M")
        file_name = (
            "LloydGeorgeSummary_"
            + ods_code
            + f"_{len(nhs_numbers)}_"
            + formatted_time
            + "."
            + file_type_output
        )
        temp_file_path = os.path.join(self.temp_output_dir, file_name)
        match file_type_output:
            case FileType.CSV:
                self.create_csv_report(temp_file_path, nhs_numbers, ods_code)
            case FileType.XLSX:
                self.create_xlsx_report(temp_file_path, nhs_numbers, ods_code)
            case FileType.PDF:
                self.create_pdf_report(temp_file_path, nhs_numbers, ods_code)
            case _:
                raise OdsReportException(400, LambdaError.UnsupportedFileType)
        logger.info(
            f"Query completed. {len(nhs_numbers)} items written to {file_name}."
        )
        if upload_to_s3:
            self.save_report_to_s3(ods_code, file_name, temp_file_path)

            if create_pre_signed_url:
                return self.get_pre_signed_url(ods_code, file_name)

    def create_csv_report(self, file_name: str, nhs_numbers: set[str], ods_code: str):
        with open(file_name, "w") as f:
            f.write(
                f"Total number of patients for ODS code {ods_code}: {len(nhs_numbers)}\n"
            )
            f.write("NHS Numbers:\n")
            f.writelines(f"{nhs_number}\n" for nhs_number in nhs_numbers)

    def create_xlsx_report(self, file_name: str, nhs_numbers: set[str], ods_code: str):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = (
            f"Total number of patients for ODS code {ods_code}: {len(nhs_numbers)}\n"
        )
        ws["A2"] = "NHS Numbers:\n"
        for row in nhs_numbers:
            ws.append([row])

        wb.save(file_name)

    def create_pdf_report(self, file_name: str, nhs_numbers: set[str], ods_code: str):
        c = canvas.Canvas(file_name, pagesize=letter)
        _, height = letter
        c.setFont("Helvetica-Bold", 16)
        x = 100
        y = 700
        c.drawString(x, height - 50, f"NHS numbers within NDR for ODS code: {ods_code}")
        c.setFont("Helvetica", 12)

        c.drawString(x, y, f"Total number of patients: {len(nhs_numbers)}")
        y -= 20
        c.drawString(x, y, "NHS Numbers:")
        y -= 20
        for row in nhs_numbers:
            if y < 40:
                c.showPage()
                y = height - 50

            c.drawString(100, y, row)
            y -= 20

        c.save()

    def save_report_to_s3(self, ods_code: str, file_name: str, temp_file_path: str):
        logger.info("Uploading the csv report to S3 bucket...")
        today = datetime.now().date()
        self.s3_service.upload_file(
            s3_bucket_name=self.reports_bucket,
            file_key=f"ods-reports/{ods_code}/{today.year}/{today.month}/{today.day}/{file_name}",
            file_name=temp_file_path,
        )

    def get_pre_signed_url(self, ods_code: str, file_name: str):
        today = datetime.now().date()
        return self.s3_service.create_download_presigned_url(
            s3_bucket_name=self.reports_bucket,
            file_key=f"ods-reports/{ods_code}/{today.year}/{today.month}/{today.day}/{file_name}",
        )
