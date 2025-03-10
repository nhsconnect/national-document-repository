import os
import tempfile

import pytest
from enums.file_type import FileType
from enums.metadata_field_names import DocumentReferenceMetadataFields
from freezegun import freeze_time
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from services.ods_report_service import OdsReportService
from utils.lambda_exceptions import OdsReportException


@pytest.fixture
def ods_report_service(mocker, set_env):
    mocker.patch("services.ods_report_service.S3Service")
    mocker.patch("services.ods_report_service.DynamoDBService")
    temp_folder = tempfile.mkdtemp()
    mocker.patch.object(tempfile, "mkdtemp", return_value=temp_folder)
    service = OdsReportService()
    return service


@pytest.fixture
def mocked_context(mocker):
    mocked_context = mocker.MagicMock()
    mocked_context.authorization = {
        "selected_organisation": {"org_ods_code": "Y12345"},
        "repository_role": "GP_ADMIN",
    }
    yield mocker.patch("services.ods_report_service.request_context", mocked_context)


@pytest.fixture
def mock_create_and_save_ods_report(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "create_and_save_ods_report")


@pytest.fixture
def mock_scan_table_with_filter(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "scan_table_with_filter")


@pytest.fixture
def mock_dynamo_service_scan_table(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service.dynamo_service, "scan_whole_table")


@pytest.fixture
def mock_create_report_csv(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "create_csv_report")


@pytest.fixture
def mock_create_report_pdf(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "create_pdf_report")


@pytest.fixture
def mock_save_report_to_s3(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "save_report_to_s3")


@pytest.fixture
def mock_get_pre_signed_url(mocker, ods_report_service):
    return mocker.patch.object(ods_report_service, "get_pre_signed_url")


def test_get_nhs_numbers_by_ods(
    ods_report_service, mock_scan_table_with_filter, mock_create_and_save_ods_report
):
    mock_scan_table_with_filter.return_value = [
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS123"},
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS456"},
    ]

    ods_report_service.get_nhs_numbers_by_ods("ODS123")

    mock_scan_table_with_filter.assert_called_once_with("ODS123")
    mock_create_and_save_ods_report.assert_called_once_with(
        "ODS123", {"NHS123", "NHS456"}, False, False, "csv"
    )


def test_get_nhs_numbers_by_ods_with_temp_folder(
    ods_report_service, mock_scan_table_with_filter, mock_create_and_save_ods_report
):
    mock_scan_table_with_filter.return_value = [
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS123"},
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS456"},
    ]

    ods_report_service.get_nhs_numbers_by_ods("ODS123", is_upload_to_s3_needed=True)

    mock_scan_table_with_filter.assert_called_once_with("ODS123")
    mock_create_and_save_ods_report.assert_called_once_with(
        "ODS123", {"NHS123", "NHS456"}, False, True, "csv"
    )
    assert ods_report_service.temp_output_dir != ""


def test_scan_table_with_filter(
    ods_report_service, mocked_context, mock_dynamo_service_scan_table
):
    mock_dynamo_service_scan_table.return_value = [
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS123"},
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS456"},
        {DocumentReferenceMetadataFields.NHS_NUMBER.value: "NHS789"},
    ]

    results = ods_report_service.scan_table_with_filter("ODS123")

    assert len(results) == 3
    assert mock_dynamo_service_scan_table.call_count == 1


def test_scan_table_with_filter_no_results(
    ods_report_service, mocked_context, mock_dynamo_service_scan_table
):
    mock_dynamo_service_scan_table.return_value = []

    with pytest.raises(OdsReportException):
        ods_report_service.scan_table_with_filter("ODS123")


@freeze_time("2024-01-01T12:00:00Z")
def test_create_and_save_ods_report_create_csv(
    ods_report_service,
    mock_create_report_csv,
    mock_save_report_to_s3,
    mock_get_pre_signed_url,
):
    ods_code = "ODS123"
    nhs_numbers = {"NHS123", "NHS456"}
    file_name = "LloydGeorgeSummary_ODS123_2_2024-01-01_12-00.csv"
    temp_file_path = os.path.join(ods_report_service.temp_output_dir, file_name)

    result = ods_report_service.create_and_save_ods_report(
        ods_code, nhs_numbers, upload_to_s3=True, file_type_output=FileType.CSV
    )

    mock_create_report_csv.assert_called_once_with(
        temp_file_path, nhs_numbers, ods_code
    )
    mock_save_report_to_s3.assert_called_once_with(ods_code, file_name, temp_file_path)
    mock_get_pre_signed_url.assert_not_called()

    assert result is None


@freeze_time("2024-01-01T12:00:00Z")
def test_create_and_save_ods_report_create_pdf(
    ods_report_service,
    mock_create_report_pdf,
    mock_save_report_to_s3,
    mock_get_pre_signed_url,
):
    ods_code = "ODS123"
    nhs_numbers = {"NHS123", "NHS456"}
    file_name = "LloydGeorgeSummary_ODS123_2_2024-01-01_12-00.pdf"
    temp_file_path = os.path.join(ods_report_service.temp_output_dir, file_name)

    ods_report_service.create_and_save_ods_report(
        ods_code, nhs_numbers, upload_to_s3=True, file_type_output=FileType.PDF
    )

    mock_create_report_pdf.assert_called_once_with(
        temp_file_path, nhs_numbers, ods_code
    )
    mock_save_report_to_s3.assert_called_once_with(ods_code, file_name, temp_file_path)
    mock_get_pre_signed_url.assert_not_called()


def test_create_and_save_ods_report_send_invalid_file_type(
    ods_report_service,
    mock_create_report_pdf,
    mock_save_report_to_s3,
    mock_get_pre_signed_url,
):
    ods_code = "ODS123"
    nhs_numbers = {"NHS123", "NHS456"}
    with pytest.raises(OdsReportException):
        ods_report_service.create_and_save_ods_report(
            ods_code,
            nhs_numbers,
            upload_to_s3=True,
            create_pre_signed_url=True,
            file_type_output="invalid",
        )

    mock_create_report_pdf.assert_not_called()
    mock_save_report_to_s3.assert_not_called()
    mock_get_pre_signed_url.assert_not_called()


@freeze_time("2024-01-01T12:00:00Z")
def test_create_and_save_ods_report_with_pre_sign_url(
    ods_report_service,
    mock_create_report_csv,
    mock_save_report_to_s3,
    mock_get_pre_signed_url,
):
    ods_code = "ODS123"
    nhs_numbers = {"NHS123", "NHS456"}
    file_name = "LloydGeorgeSummary_ODS123_2_2024-01-01_12-00.csv"
    mock_pre_sign_url = "https://presigned.url"

    mock_get_pre_signed_url.return_value = mock_pre_sign_url
    temp_file_path = os.path.join(ods_report_service.temp_output_dir, file_name)

    result = ods_report_service.create_and_save_ods_report(
        ods_code, nhs_numbers, True, True
    )

    mock_create_report_csv.assert_called_once_with(
        temp_file_path, nhs_numbers, ods_code
    )
    mock_save_report_to_s3.assert_called_once_with(ods_code, file_name, temp_file_path)
    mock_get_pre_signed_url.assert_called_once_with(ods_code, file_name)
    assert result == mock_pre_sign_url


def test_create_report_csv(ods_report_service, tmp_path):
    nhs_numbers = {"NHS123", "NHS456"}
    file_name = tmp_path / "test_report.csv"
    ods_code = "ODS123"

    ods_report_service.create_csv_report(str(file_name), nhs_numbers, ods_code)

    with open(file_name, "r") as f:
        content = f.readlines()

    assert (
        f"Total number of patients for ODS code {ods_code}: {len(nhs_numbers)}\n"
        in content
    )
    assert "NHS Numbers:\n" in content
    assert "NHS123\n" in content
    assert "NHS456\n" in content


def test_create_xlsx_report(ods_report_service, tmp_path):
    file_name = "test_report.xlsx"
    nhs_numbers = {"NHS123456", "NHS654321", "NHS111222"}
    ods_code = "ODS123"

    ods_report_service.create_xlsx_report(file_name, nhs_numbers, ods_code)

    assert os.path.exists(file_name)

    wb = load_workbook(file_name)
    ws = wb.active

    assert (
        ws["A1"].value
        == f"Total number of patients for ODS code {ods_code}: {len(nhs_numbers)}\n"
    )
    assert ws["A2"].value == "NHS Numbers:\n"

    for i, nhs_number in enumerate(nhs_numbers, start=3):  # Start from row 3
        assert ws.cell(row=i, column=1).value == nhs_number

    os.remove(file_name)


@freeze_time("2024-01-01T12:00:00Z")
def test_create_pdf_report(ods_report_service):
    file_name = "test_report.pdf"
    nhs_numbers = {"NHS123456", "NHS654321", "NHS111222"}
    ods_code = "ODS123"

    ods_report_service.create_pdf_report(file_name, nhs_numbers, ods_code)

    assert os.path.exists(file_name)

    reader = PdfReader(file_name)
    assert len(reader.pages) > 0

    first_page = reader.pages[0].extract_text()

    assert f"NHS numbers within NDR for ODS code: {ods_code}" in first_page
    assert f"Total number of patients: {len(nhs_numbers)}" in first_page
    for nhs_number in nhs_numbers:
        assert nhs_number in first_page

    os.remove(file_name)


@freeze_time("2024-01-01T12:00:00Z")
def test_save_report_to_s3(ods_report_service, mocker):
    mocker.patch.object(ods_report_service.s3_service, "upload_file")

    ods_report_service.save_report_to_s3("ODS123", "test_report.csv", "path/to/file")

    ods_report_service.s3_service.upload_file.assert_called_once_with(
        s3_bucket_name="test_statistics_report_bucket",
        file_key="ods-reports/ODS123/2024/1/1/test_report.csv",
        file_name="path/to/file",
    )


@freeze_time("2024-01-01T12:00:00Z")
def test_get_pre_signed_url(ods_report_service, mocker):
    mocker.patch.object(ods_report_service.s3_service, "create_download_presigned_url")

    ods_report_service.get_pre_signed_url("ODS123", "test_report.csv")

    ods_report_service.s3_service.create_download_presigned_url.assert_called_once_with(
        s3_bucket_name="test_statistics_report_bucket",
        file_key="ods-reports/ODS123/2024/1/1/test_report.csv",
    )
